from openpyxl import load_workbook
wb=load_workbook(r"C:\Users\Gautam G\OneDrive\Desktop\students.xlsx")
sheet=wb.active
for row in sheet.iter_rows(min_row=2,values_only=True):
    name,marks=row
    highest=0
    if marks>highest:
        highest=marks   
print(f"The highest marks are {highest} obtained by {name}.")